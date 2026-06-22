"""
Import products from KrissMaagiic_Products.xlsx (Desktop) → MongoDB.
Run: python scripts/import-products-excel.py
     python scripts/import-products-excel.py --dry-run   (preview only, no writes)
"""

import sys, os, json, re
from pathlib import Path

# ── Load .env.local ──────────────────────────────────────────────────────────
env_path = Path(__file__).parent.parent / '.env.local'
if env_path.exists():
    for line in env_path.read_text(encoding='utf-8').splitlines():
        if '=' in line and not line.startswith('#'):
            k, _, v = line.partition('=')
            os.environ.setdefault(k.strip(), v.strip())

MONGO_URI = os.environ.get('MONGODB_URI')
if not MONGO_URI:
    sys.exit('ERROR: MONGODB_URI not found in .env.local')

EXCEL_PATH = Path.home() / 'Desktop' / 'KrissMaagiic_Products.xlsx'
if not EXCEL_PATH.exists():
    sys.exit(f'ERROR: File not found: {EXCEL_PATH}')

DRY_RUN = '--dry-run' in sys.argv

# ── Open workbook ─────────────────────────────────────────────────────────────
import openpyxl
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# Column index map (1-based, matching the generated headers)
COL = {
    'id': 1, 'name': 2, 'category': 3, 'subcategory': 4,
    'price': 5, 'originalPrice': 6, 'usdPrice': 7, 'originalUsdPrice': 8,
    'badge': 9, 'stock': 10, 'desc': 11,
    'purpose': 12, 'crystalsIncluded': 13, 'associatedChakras': 14,
    'longDescText': 15, 'benefits': 16,
    'recommendedHand': 17, 'whenToWear': 18, 'howToEnergize': 19,
    'affirmation': 20, 'careInstructions': 21, 'disclaimer': 22,
    'chakras': 23, 'image': 24, 'images': 25, 'variants': 26,
}

VALID_BADGES = {'Popular', 'New', 'Sale', 'Bestseller'}

def cell(row, col_name):
    v = row[COL[col_name] - 1]
    return str(v).strip() if v is not None and str(v).strip() not in ('None', '') else ''

def num(row, col_name, default=None):
    v = row[COL[col_name] - 1]
    try:
        return float(v) if v is not None and str(v).strip() not in ('', 'None') else default
    except (ValueError, TypeError):
        return default

def split_pipe(s):
    return [x.strip() for x in s.split('|') if x.strip()] if s else []

def split_comma(s):
    return [x.strip() for x in s.split(',') if x.strip()] if s else []

def build_long_desc(row):
    d = {}
    for key, col in [
        ('purpose', 'purpose'), ('crystalsIncluded', 'crystalsIncluded'),
        ('associatedChakras', 'associatedChakras'), ('description', 'longDescText'),
        ('recommendedHand', 'recommendedHand'), ('whenToWear', 'whenToWear'),
        ('howToEnergize', 'howToEnergize'), ('affirmation', 'affirmation'),
        ('disclaimer', 'disclaimer'),
    ]:
        v = cell(row, col)
        if v:
            d[key] = v
    benefits = split_pipe(cell(row, 'benefits'))
    if benefits:
        d['benefits'] = benefits
    care = split_pipe(cell(row, 'careInstructions'))
    if care:
        d['careInstructions'] = care
    return json.dumps(d, ensure_ascii=False) if d else ''

# ── Parse all rows ────────────────────────────────────────────────────────────
docs = []
skipped = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        continue
    # rows[0] = header, rows[1:] = data
    for row in rows[1:]:
        slug = cell(row, 'id')
        name = cell(row, 'name')
        if not slug or not name:
            continue

        price = num(row, 'price')
        if price is None:
            skipped.append(f'{slug}: missing price')
            continue

        desc = cell(row, 'desc')
        if not desc:
            skipped.append(f'{slug}: missing desc')
            continue

        image = cell(row, 'image')
        if not image:
            skipped.append(f'{slug}: missing image')
            continue

        badge_raw = cell(row, 'badge')
        badge = badge_raw if badge_raw in VALID_BADGES else None

        orig_price = num(row, 'originalPrice')
        usd = num(row, 'usdPrice', 0)
        orig_usd = num(row, 'originalUsdPrice')
        stock = int(num(row, 'stock', 99) or 99)

        images_str = cell(row, 'images')
        images = split_pipe(images_str) if images_str else [image]

        chakras = split_comma(cell(row, 'chakras'))

        variants_raw = cell(row, 'variants')
        variants = []
        if variants_raw:
            try:
                variants = json.loads(variants_raw)
            except Exception:
                pass

        long_desc = build_long_desc(row)

        doc = {
            'slug': slug,
            'name': name,
            'category': cell(row, 'category') or sheet_name,
            'subcategory': cell(row, 'subcategory'),
            'price': price,
            'originalPrice': orig_price,
            'image': image,
            'images': images,
            'usdPrice': usd or 0,
            'originalUsdPrice': orig_usd,
            'badge': badge,
            'desc': desc,
            'longDesc': long_desc,
            'chakras': chakras,
            'stock': stock,
            'variants': variants if variants else None,
            'active': True,
            'isDeleted': False,
        }
        docs.append(doc)

wb.close()

print(f'\nParsed {len(docs)} products from {len(wb.sheetnames)} sheets')
if skipped:
    print(f'Skipped {len(skipped)} rows:')
    for s in skipped:
        print(f'  - {s}')

if DRY_RUN:
    print('\n-- DRY RUN: no changes written --')
    for d in docs[:5]:
        print(f"  {d['slug']} | {d['name']} | INR {d['price']} | stock={d['stock']}")
    print(f'  ... and {len(docs)-5} more')
    sys.exit(0)

# ── Upsert to MongoDB ─────────────────────────────────────────────────────────
from pymongo import MongoClient, UpdateOne
from urllib.parse import urlparse

client = MongoClient(MONGO_URI)
db_name = urlparse(MONGO_URI).path.lstrip('/')
if '?' in db_name:
    db_name = db_name.split('?')[0]
db = client[db_name]
collection = db['products']

ops = []
for doc in docs:
    slug = doc.pop('slug')
    ops.append(UpdateOne(
        {'slug': slug},
        {'$set': doc, '$setOnInsert': {'slug': slug}},
        upsert=True,
    ))

result = collection.bulk_write(ops, ordered=False)
client.close()

print(f'\nDone!')
print(f'  Inserted : {result.upserted_count}')
print(f'  Updated  : {result.modified_count}')
print(f'  Matched  : {result.matched_count}')
